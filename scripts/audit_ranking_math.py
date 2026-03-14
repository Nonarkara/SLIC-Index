from __future__ import annotations

import ast
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from verified_source_pipeline import (
    compute_ranked_rows,
    prepare_verified_source_pack,
    source_pack_completion_summary,
    validate_source_pack,
)

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "output" / "spreadsheet" / "slic_scoring_workbook.xlsx"
METHODOLOGY_DATA_PATH = ROOT / "src" / "methodologyData.ts"
METHODOLOGY_PDF_PATH = ROOT / "scripts" / "generate_methodology_pdf.py"
RANKINGS_DATA_PATH = ROOT / "src" / "rankingsData.ts"
SITE_COPY_PATH = ROOT / "src" / "siteCopy.ts"
SITE_FOOTER_PATH = ROOT / "src" / "SiteFooter.tsx"
VALIDATOR_PATH = ROOT / "scripts" / "validate_slic_workbook.py"


def extract_workbook_weights() -> dict[str, int]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=False, read_only=True)
    scores_sheet = workbook["Scores"]
    headers = [cell.value for cell in next(scores_sheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: index for index, header in enumerate(headers)}
    row = next(scores_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    formula = row[header_index["slic_score"]]
    if not isinstance(formula, str):
        raise RuntimeError("Scores.slic_score is missing its formula")

    pattern = re.compile(
        r'IF\((?P<ref>[A-Z]+2)="",0,(?P=ref)\*(?P<weight>\d+)\)',
    )
    weights = [int(match.group("weight")) for match in pattern.finditer(formula)]
    if len(weights) != 5:
        raise RuntimeError(f"Could not parse workbook pillar weights from formula: {formula}")

    return {
        "pressure": weights[0],
        "viability": weights[1],
        "capability": weights[2],
        "community": weights[3],
        "creative": weights[4],
    }


def extract_formula_strings(path: Path) -> set[str]:
    text = path.read_text(encoding="utf-8")
    if path.suffix == ".py":
        tree = ast.parse(text)
        formulas = {
            re.sub(r"\s+", " ", node.value).strip()
            for node in ast.walk(tree)
            if isinstance(node, ast.Constant)
            and isinstance(node.value, str)
            and "SLIC(c) =" in node.value
        }
        return formulas

    matches = re.findall(r"SLIC\(c\)\s*=\s*[^\"]+", text)
    return {re.sub(r"\s+", " ", match).strip() for match in matches}


def parse_formula_weights(formula: str) -> dict[str, float]:
    pattern = re.compile(r"([0-9.]+)\s+([A-Za-z]+)")
    weights: dict[str, float] = {}
    for raw_weight, raw_label in pattern.findall(formula):
        label = raw_label.lower()
        if label.startswith("daily"):
            label = "viability"
        elif label.startswith("safety"):
            label = "viability"
        elif label.startswith("human"):
            label = "capability"
        elif label.startswith("community"):
            label = "community"
        elif label.startswith("business"):
            label = "creative"
        elif label.startswith("creative"):
            label = "creative"
        elif label.startswith("pressure"):
            label = "pressure"
        weights[label] = float(raw_weight)
    return weights


def workbook_is_unpopulated() -> tuple[bool, str]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    country_context = workbook["Country_Context"]
    city_inputs = workbook["City_Inputs"]
    leaderboard = workbook["Leaderboard"]

    country_headers = [cell.value for cell in next(country_context.iter_rows(min_row=1, max_row=1))]
    country_index = {header: index for index, header in enumerate(country_headers)}
    context_fields = [
        "gdp_per_capita_ppp",
        "gdp_growth",
        "gini_coefficient",
        "ppp_private_consumption",
        "tax_rate_assumption",
        "household_debt_proxy",
    ]
    populated_context = 0
    sampled_countries = 0
    for row in country_context.iter_rows(min_row=2, values_only=True):
        sampled_countries += 1
        populated_context += sum(
            row[country_index[field]] not in ("", None) for field in context_fields
        )

    city_headers = [cell.value for cell in next(city_inputs.iter_rows(min_row=1, max_row=1))]
    city_index = {header: index for index, header in enumerate(city_headers)}
    raw_fields = [
        "gross_income",
        "rent",
        "utilities",
        "transit_cost",
        "internet_cost",
        "food_cost",
        "personal_safety_raw",
        "healthcare_quality_raw",
        "entrepreneurial_dynamism_raw",
    ]
    populated_city_inputs = 0
    sampled_cities = 0
    for row in city_inputs.iter_rows(min_row=2, values_only=True):
        sampled_cities += 1
        populated_city_inputs += sum(
            row[city_index[field]] not in ("", None) for field in raw_fields
        )

    leaderboard_headers = [cell.value for cell in next(leaderboard.iter_rows(min_row=2, max_row=2))]
    leaderboard_index = {header: index for index, header in enumerate(leaderboard_headers)}
    leaderboard_values = 0
    for row in leaderboard.iter_rows(min_row=3, max_row=12, values_only=True):
        leaderboard_values += sum(
            row[leaderboard_index[field]] not in ("", None)
            for field in ("rank", "display_name", "slic_score")
        )

    empty = populated_context == 0 and populated_city_inputs == 0 and leaderboard_values == 0
    detail = (
        f"country inputs populated={populated_context}/{sampled_countries * len(context_fields)}, "
        f"city inputs populated={populated_city_inputs}/{sampled_cities * len(raw_fields)}, "
        f"leaderboard values populated={leaderboard_values}/30"
    )
    return empty, detail


def source_lineage_is_ready() -> tuple[bool, str]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    details: list[str] = []

    country_context = workbook["Country_Context"]
    country_headers = [cell.value for cell in next(country_context.iter_rows(min_row=1, max_row=1))]
    country_index = {header: index for index, header in enumerate(country_headers)}
    country_rows = list(country_context.iter_rows(min_row=2, values_only=True))
    country_url_count = sum(
        row[country_index["source_url"]] not in ("", None) for row in country_rows
    )
    country_tier_count = sum(
        row[country_index["source_tier"]] not in ("", None) for row in country_rows
    )
    details.append(
        f"Country_Context source_url={country_url_count}/{len(country_rows)}, "
        f"source_tier={country_tier_count}/{len(country_rows)}"
    )

    city_inputs = workbook["City_Inputs"]
    city_headers = [cell.value for cell in next(city_inputs.iter_rows(min_row=1, max_row=1))]
    city_index = {header: index for index, header in enumerate(city_headers)}
    source_url_headers = [header for header in city_headers if isinstance(header, str) and header.endswith("_source_url")]
    source_tier_headers = [header for header in city_headers if isinstance(header, str) and header.endswith("_source_tier")]
    city_rows = list(city_inputs.iter_rows(min_row=2, values_only=True))
    city_url_slots = len(city_rows) * len(source_url_headers)
    city_tier_slots = len(city_rows) * len(source_tier_headers)
    city_url_count = sum(
        row[city_index[header]] not in ("", None)
        for row in city_rows
        for header in source_url_headers
    )
    city_tier_count = sum(
        row[city_index[header]] not in ("", None)
        for row in city_rows
        for header in source_tier_headers
    )
    details.append(
        f"City_Inputs source_url={city_url_count}/{city_url_slots}, "
        f"source_tier={city_tier_count}/{city_tier_slots}"
    )

    ready = country_url_count == len(country_rows) and city_url_count == city_url_slots
    return ready, "; ".join(details)


def frontend_uses_synthetic_generation() -> bool:
    text = RANKINGS_DATA_PATH.read_text(encoding="utf-8")
    return "export const globalRankings: FullRankedCity[] = rankingPublication.publishable ? buildGlobalRankings() : [];" in text


def official_copy_claims_declared_model() -> list[Path]:
    paths: list[Path] = []
    phrase = "generated by the declared SLIC"
    for path in (SITE_COPY_PATH, SITE_FOOTER_PATH):
        if phrase in path.read_text(encoding="utf-8"):
            paths.append(path)
    return paths


def run_validator() -> tuple[bool, str]:
    result = subprocess.run(
        [sys.executable, str(VALIDATOR_PATH)],
        capture_output=True,
        text=True,
        check=False,
        cwd=ROOT,
    )
    output = (result.stdout + result.stderr).strip()
    return result.returncode == 0, output or "validator produced no output"


def main() -> int:
    if not WORKBOOK_PATH.exists():
        print(f"Missing workbook: {WORKBOOK_PATH}")
        return 1

    findings: list[str] = []

    validator_ok, validator_output = run_validator()
    if not validator_ok:
        findings.append(f"Workbook validation failed: {validator_output}")

    prepare_verified_source_pack(force=False)
    validation = validate_source_pack()
    ranked_rows = compute_ranked_rows(validation) if not validation.issues else []

    workbook_weights = extract_workbook_weights()
    canonical_formula = (
        "SLIC(c) = "
        f"{workbook_weights['pressure'] / 100:.2f} Pressure + "
        f"{workbook_weights['viability'] / 100:.2f} Viability + "
        f"{workbook_weights['capability'] / 100:.2f} Capability + "
        f"{workbook_weights['community'] / 100:.2f} Community + "
        f"{workbook_weights['creative'] / 100:.2f} Creative"
    )

    for path in (METHODOLOGY_DATA_PATH, METHODOLOGY_PDF_PATH):
        formulas = extract_formula_strings(path)
        mismatches = []
        for formula in sorted(formulas):
            parsed = parse_formula_weights(formula)
            if not parsed:
                continue
            expected = {
                "pressure": workbook_weights["pressure"] / 100,
                "viability": workbook_weights["viability"] / 100,
                "capability": workbook_weights["capability"] / 100,
                "community": workbook_weights["community"] / 100,
                "creative": workbook_weights["creative"] / 100,
            }
            if parsed != expected:
                mismatches.append(formula)
        if mismatches:
            findings.append(
                f"{path.relative_to(ROOT)} declares formula(s) that do not match workbook weights {canonical_formula}: "
                + " | ".join(mismatches)
            )

    if validation.issues:
        findings.extend(validation.issues)

    if not ranked_rows:
        findings.append(
            "Verified source pack does not currently produce any ranked city rows: "
            + source_pack_completion_summary(validation)
        )

    if frontend_uses_synthetic_generation():
        findings.append(
            "src/rankingsData.ts still routes the public ranking board through the synthetic preview engine instead of the verified export."
        )

    misleading_copy_paths = official_copy_claims_declared_model()
    if misleading_copy_paths and (frontend_uses_synthetic_generation() or not ranked_rows):
        findings.append(
            "Public copy claims the declared SLIC model is driving the ordering while the current site ranking path is synthetic: "
            + ", ".join(str(path.relative_to(ROOT)) for path in misleading_copy_paths)
        )

    print("Ranking math audit")
    print(f"- Workbook validator: {'PASS' if validator_ok else 'FAIL'}")
    print(f"- Workbook canonical weights: {canonical_formula}")
    print(f"- Source-pack completion: {source_pack_completion_summary(validation)}")
    print(f"- Ranked rows from verified inputs: {len(ranked_rows)}")

    if findings:
        print("- Findings:")
        for finding in findings:
            print(f"  * {finding}")
        return 1

    print("- Findings: none")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
