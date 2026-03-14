from __future__ import annotations

from openpyxl import load_workbook

from extract_economist_reference import build_reference_csv
from generate_google_sheets_template import (
    BOLD_FONT,
    HEADER_FILL,
    build_coverage_sheet,
    build_guide_sheet,
    write_sheet_table,
)
from generate_slic_workbook import WORKBOOK_PATH, build_city_universe, main as build_workbook, unique_countries
from source_guides import CITY_SOURCE_PLAYBOOK_PATH, FIELD_SOURCE_GUIDE_PATH, ensure_source_guide_files
from verified_source_pipeline import (
    CITY_INPUTS_PATH,
    COUNTRY_CONTEXT_PATH,
    PROVIDERS_PATH,
    ROOT,
    apply_source_pack_to_workbook,
    prepare_verified_source_pack,
    read_csv_rows,
    source_pack_completion_summary,
    validate_source_pack,
)


OUTPUT_PATH = ROOT / "output" / "spreadsheet" / "slic_internal_analyst_workbook.xlsx"
COUNTRY_ALIASES = {
    "usa": "unitedstates",
    "us": "unitedstates",
    "uk": "unitedkingdom",
    "uae": "unitedarabemirates",
}


def normalize_key(value: str) -> str:
    compact = "".join(character for character in value.lower() if character.isalnum())
    return COUNTRY_ALIASES.get(compact, compact)


def append_internal_reference_notes(workbook) -> None:
    sheet = workbook["Google_Sheets_Guide"]
    start_row = sheet.max_row + 2
    sheet.cell(row=start_row, column=1, value="Internal reference layer").font = BOLD_FONT
    sheet.cell(row=start_row, column=1).fill = HEADER_FILL

    rows = [
        (
            "Economist_Reference",
            "Auto-extracted from The World in Numbers PDF for analyst speed only. Do not treat these values as publishable scoring inputs.",
        ),
        (
            "Economist_Match_Review",
            "One review row per SLIC country. Promote nothing automatically into Country_Context; every value still needs a trusted non-reference source.",
        ),
        (
            "Reference-only provider rule",
            "economist_world_in_numbers is Tier 4 and reference_only=true, so it cannot make the public ranking publishable on its own.",
        ),
    ]

    for offset, row in enumerate(rows, start=1):
        sheet.cell(row=start_row + offset, column=1, value=row[0]).font = BOLD_FONT
        sheet.cell(row=start_row + offset, column=1).fill = HEADER_FILL
        sheet.cell(row=start_row + offset, column=2, value=row[1])


def build_match_review_rows(reference_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    reference_by_country = {
        normalize_key(row["country"]): row
        for row in reference_rows
    }

    rows: list[dict[str, str]] = []
    for country in unique_countries(build_city_universe()):
        matched = reference_by_country.get(normalize_key(country))
        rows.append(
            {
                "slic_country": country,
                "match_status": "matched" if matched else "unmatched",
                "matched_economist_country": matched["country"] if matched else "",
                "page_number": matched["page_number"] if matched else "",
                "gdp_growth_pct": matched["gdp_growth_pct"] if matched else "",
                "gdp_per_person_usd": matched["gdp_per_person_usd"] if matched else "",
                "gdp_per_person_ppp_usd": matched["gdp_per_person_ppp_usd"] if matched else "",
                "inflation_pct": matched["inflation_pct"] if matched else "",
                "budget_balance_pct_gdp": matched["budget_balance_pct_gdp"] if matched else "",
                "population_millions": matched["population_millions"] if matched else "",
                "parse_confidence": matched["parse_confidence"] if matched else "",
                "review_status": matched["review_status"] if matched else "needs_match",
                "source_label": matched["source_label"] if matched else "",
                "source_date": matched["source_date"] if matched else "",
                "promote_to_country_context": "",
                "analyst_notes": "",
            }
        )
    return rows


def main() -> int:
    build_workbook()
    prepare_verified_source_pack(force=False)
    ensure_source_guide_files(force=False)
    validation = validate_source_pack()
    if not validation.issues:
        apply_source_pack_to_workbook(validation)

    _, reference_rows = build_reference_csv()
    match_review_rows = build_match_review_rows(reference_rows)

    workbook = load_workbook(WORKBOOK_PATH)
    build_guide_sheet(workbook, validation.issues, source_pack_completion_summary(validation))
    build_coverage_sheet(workbook, validation)
    append_internal_reference_notes(workbook)
    write_sheet_table(workbook, "Providers_Registry", read_csv_rows(PROVIDERS_PATH))
    write_sheet_table(workbook, "Field_Source_Guide", read_csv_rows(FIELD_SOURCE_GUIDE_PATH))
    write_sheet_table(workbook, "City_Source_Playbook", read_csv_rows(CITY_SOURCE_PLAYBOOK_PATH))
    write_sheet_table(workbook, "Country_Source_Rows", read_csv_rows(COUNTRY_CONTEXT_PATH))
    write_sheet_table(workbook, "City_Source_Rows", read_csv_rows(CITY_INPUTS_PATH))
    write_sheet_table(workbook, "Economist_Reference", reference_rows)
    write_sheet_table(workbook, "Economist_Match_Review", match_review_rows)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)

    matched_rows = sum(1 for row in match_review_rows if row["match_status"] == "matched")

    print(f"Wrote {OUTPUT_PATH.relative_to(ROOT)}")
    print(f"- Economist reference rows: {len(reference_rows)}")
    print(f"- SLIC country matches: {matched_rows}/{len(unique_countries(build_city_universe()))}")
    print(f"- Source-pack completion: {source_pack_completion_summary(validation)}")
    if validation.issues:
        for issue in validation.issues:
            print(f"- {issue}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
