from __future__ import annotations

from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from generate_slic_workbook import WORKBOOK_PATH, main as build_workbook
from source_guides import CITY_SOURCE_PLAYBOOK_PATH, FIELD_SOURCE_GUIDE_PATH, ensure_source_guide_files
from verified_source_pipeline import (
    CITY_FIELD_SPECS,
    CITY_INPUTS_PATH,
    COUNTRY_FIELD_SPECS,
    COUNTRY_CONTEXT_PATH,
    PROVIDERS_PATH,
    ROOT,
    apply_source_pack_to_workbook,
    build_city_integrity_watchlist,
    build_integrity_dashboard_rows,
    prepare_verified_source_pack,
    read_csv_rows,
    source_pack_completion_summary,
    validate_source_pack,
)


OUTPUT_PATH = ROOT / "output" / "spreadsheet" / "slic_google_sheets_template.xlsx"
PUBLIC_PATH = ROOT / "public" / "downloads" / "slic-google-sheets-template.xlsx"

TITLE_FILL = PatternFill("solid", fgColor="0F2B60")
HEADER_FILL = PatternFill("solid", fgColor="EAF2FF")
SUBHEADER_FILL = PatternFill("solid", fgColor="F7FAFC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


def write_sheet_table(workbook, title: str, rows: list[dict[str, str]]) -> None:
    if title in workbook.sheetnames:
        del workbook[title]

    sheet = workbook.create_sheet(title)
    if not rows:
        sheet["A1"] = "No rows available"
        return

    headers = list(rows[0].keys())
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = BOLD_FONT

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_index, column=column_index, value=row[header])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells[: min(len(column_cells), 40)])
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 40)


def build_guide_sheet(workbook, validation_issues: list[str], completion_summary: str) -> None:
    if "Google_Sheets_Guide" in workbook.sheetnames:
        del workbook["Google_Sheets_Guide"]

    sheet = workbook.create_sheet("Google_Sheets_Guide", 0)
    rows = [
        ("SLIC Google Sheets Template",),
        (
            "Import this workbook into Google Sheets to collect trusted source inputs, inspect formulas, and rerank transparently.",
        ),
        ("Current verified-source completion", completion_summary),
        ("What to edit", "Country_Context and City_Inputs"),
        (
            "What not to overwrite",
            "Keep formulas intact in Norm_Stats, Scores, and Leaderboard so the board recalculates correctly.",
        ),
        (
            "Trusted source policy",
            "Every filled value should keep a visible source URL, source tier, source date, reference period, source scope, proxy status, and note trail.",
        ),
        (
            "Provider registry",
            "Use Providers_Registry as the starter reference for approved source families and tiers.",
        ),
        (
            "Integrity metadata",
            "Use reference_period for the measurement window, source_scope for city/metro/regional/national/international coverage, and proxy_status to distinguish direct evidence from approved proxies.",
        ),
        (
            "Raw audit tabs",
            "Country_Source_Rows and City_Source_Rows mirror the verified-source CSV pack so the collection workflow stays auditable.",
        ),
        (
            "Source guide tabs",
            "Field_Source_Guide sets the preferred evidence for each city metric, and City_Source_Playbook adds curated city-specific starting points such as Milan and Bologna.",
        ),
        (
            "Publishing rule",
            "Only publish rankings after the verified source pack passes integrity validation, freshness checks, and workbook audit rules.",
        ),
    ]

    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = WHITE_FONT
    sheet["A1"].parent.merge_cells("A1:B1")
    for row_index in range(2, len(rows) + 1):
        sheet.cell(row=row_index, column=1).font = BOLD_FONT
        sheet.cell(row=row_index, column=1).fill = SUBHEADER_FILL

    issue_start = len(rows) + 2
    sheet.cell(row=issue_start, column=1, value="Current validation issues").font = BOLD_FONT
    sheet.cell(row=issue_start, column=1).fill = HEADER_FILL
    if validation_issues:
        for offset, issue in enumerate(validation_issues, start=1):
            sheet.cell(row=issue_start + offset, column=1, value=f"- {issue}")
    else:
        sheet.cell(row=issue_start + 1, column=1, value="No source-pack validation issues currently detected.")

    workflow_start = issue_start + len(validation_issues) + 3
    sheet.cell(row=workflow_start, column=1, value="Suggested workflow").font = BOLD_FONT
    sheet.cell(row=workflow_start, column=1).fill = HEADER_FILL
    workflow_steps = [
        "1. Fill country context rows first from official sources.",
        "2. Use Field_Source_Guide and City_Source_Playbook before touching City_Source_Rows.",
        "3. Fill city inputs with source URLs, source dates, reference periods, source scope, and proxy status attached to each row.",
        "4. Review Data_Quality and Integrity_Watchlist after each sourcing pass.",
        "5. Re-import or keep editing in Google Sheets until the Leaderboard starts producing ranked rows with no integrity findings.",
        "6. Sync the same values back into the verified-source CSV pack for build-time publication.",
    ]
    for offset, step in enumerate(workflow_steps, start=1):
        sheet.cell(row=workflow_start + offset, column=1, value=step)

    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 36
    sheet.column_dimensions["B"].width = 112


def build_coverage_sheet(workbook, validation) -> None:
    if "Source_Coverage" in workbook.sheetnames:
        del workbook["Source_Coverage"]

    sheet = workbook.create_sheet("Source_Coverage", 1)
    country_rows = read_csv_rows(COUNTRY_CONTEXT_PATH)
    city_rows = read_csv_rows(CITY_INPUTS_PATH)
    playbook_rows = read_csv_rows(CITY_SOURCE_PLAYBOOK_PATH)

    def filled_count(rows: list[dict[str, str]], key: str, value: str) -> int:
        return sum(1 for row in rows if row.get(key, "") == value and row.get("value", "").strip())

    sheet["A1"] = "Source Coverage Dashboard"
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = WHITE_FONT
    sheet.merge_cells("A1:D1")

    summary_rows = [
        ("Country_Context completion", validation.stats["country_values"], validation.stats["country_slots"]),
        ("City_Inputs completion", validation.stats["city_values"], validation.stats["city_slots"]),
        ("Curated city playbook rows", len(playbook_rows), len(playbook_rows)),
        ("Current city universe", validation.stats["city_count"], validation.stats["city_count"]),
    ]
    sheet["A3"] = "Overall"
    sheet["A3"].fill = HEADER_FILL
    sheet["A3"].font = BOLD_FONT
    for column_index, header in enumerate(("Metric", "Filled", "Total", "Completion"), start=1):
        cell = sheet.cell(row=4, column=column_index, value=header)
        cell.fill = SUBHEADER_FILL
        cell.font = BOLD_FONT
    for row_index, (label, filled, total) in enumerate(summary_rows, start=5):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=filled)
        sheet.cell(row=row_index, column=3, value=total)
        completion = f"{(filled / total * 100):.1f}%" if total else "0.0%"
        sheet.cell(row=row_index, column=4, value=completion)

    section_row = 11
    sheet.cell(row=section_row, column=1, value="Country Field Coverage").fill = HEADER_FILL
    sheet.cell(row=section_row, column=1).font = BOLD_FONT
    for column_index, header in enumerate(("Field", "Filled", "Total", "Completion"), start=1):
        cell = sheet.cell(row=section_row + 1, column=column_index, value=header)
        cell.fill = SUBHEADER_FILL
        cell.font = BOLD_FONT
    for offset, spec in enumerate(COUNTRY_FIELD_SPECS, start=2):
        filled = filled_count(country_rows, "field", spec["field"])
        total = validation.stats["country_count"]
        sheet.cell(row=section_row + offset, column=1, value=spec["label"])
        sheet.cell(row=section_row + offset, column=2, value=filled)
        sheet.cell(row=section_row + offset, column=3, value=total)
        sheet.cell(row=section_row + offset, column=4, value=f"{(filled / total * 100):.1f}%")

    section_row = 20
    sheet.cell(row=section_row, column=1, value="Italy Spotlight").fill = HEADER_FILL
    sheet.cell(row=section_row, column=1).font = BOLD_FONT
    for column_index, header in enumerate(("Target", "Filled", "Total", "Completion"), start=1):
        cell = sheet.cell(row=section_row + 1, column=column_index, value=header)
        cell.fill = SUBHEADER_FILL
        cell.font = BOLD_FONT

    italy_country_rows = [row for row in country_rows if row.get("country", "") == "Italy"]
    italy_country_filled = sum(1 for row in italy_country_rows if row.get("value", "").strip())
    spotlight_rows = [
        ("Italy Country_Context", italy_country_filled, len(italy_country_rows)),
    ]
    for city_id, label in (("it-milan", "Milan City_Inputs"), ("it-bologna", "Bologna City_Inputs")):
        matched_rows = [row for row in city_rows if row.get("city_id", "") == city_id]
        filled = sum(1 for row in matched_rows if row.get("value", "").strip())
        spotlight_rows.append((label, filled, len(matched_rows)))

    for offset, (label, filled, total) in enumerate(spotlight_rows, start=2):
        sheet.cell(row=section_row + offset, column=1, value=label)
        sheet.cell(row=section_row + offset, column=2, value=filled)
        sheet.cell(row=section_row + offset, column=3, value=total)
        sheet.cell(row=section_row + offset, column=4, value=f"{(filled / total * 100):.1f}%" if total else "0.0%")

    section_row = 26
    sheet.cell(row=section_row, column=1, value="City Field Coverage").fill = HEADER_FILL
    sheet.cell(row=section_row, column=1).font = BOLD_FONT
    for column_index, header in enumerate(("Field", "Filled", "Total", "Completion"), start=1):
        cell = sheet.cell(row=section_row + 1, column=column_index, value=header)
        cell.fill = SUBHEADER_FILL
        cell.font = BOLD_FONT
    for offset, spec in enumerate(CITY_FIELD_SPECS, start=2):
        filled = filled_count(city_rows, "field", spec["field"])
        total = validation.stats["city_count"]
        sheet.cell(row=section_row + offset, column=1, value=spec["label"])
        sheet.cell(row=section_row + offset, column=2, value=filled)
        sheet.cell(row=section_row + offset, column=3, value=total)
        sheet.cell(row=section_row + offset, column=4, value=f"{(filled / total * 100):.1f}%")

    sheet.freeze_panes = "A4"
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 14


def main() -> int:
    build_workbook()
    prepare_verified_source_pack(force=False)
    ensure_source_guide_files(force=False)
    validation = validate_source_pack()
    if not validation.issues:
        apply_source_pack_to_workbook(validation)

    workbook = load_workbook(WORKBOOK_PATH)
    build_guide_sheet(workbook, validation.issues, source_pack_completion_summary(validation))
    build_coverage_sheet(workbook, validation)
    write_sheet_table(workbook, "Data_Quality", build_integrity_dashboard_rows(validation))
    write_sheet_table(workbook, "Integrity_Watchlist", build_city_integrity_watchlist(validation))
    write_sheet_table(workbook, "Providers_Registry", read_csv_rows(PROVIDERS_PATH))
    write_sheet_table(workbook, "Field_Source_Guide", read_csv_rows(FIELD_SOURCE_GUIDE_PATH))
    write_sheet_table(workbook, "City_Source_Playbook", read_csv_rows(CITY_SOURCE_PLAYBOOK_PATH))
    write_sheet_table(workbook, "Country_Source_Rows", read_csv_rows(COUNTRY_CONTEXT_PATH))
    write_sheet_table(workbook, "City_Source_Rows", read_csv_rows(CITY_INPUTS_PATH))

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    shutil.copy2(OUTPUT_PATH, PUBLIC_PATH)

    print(f"Wrote {OUTPUT_PATH.relative_to(ROOT)}")
    print(f"Copied {PUBLIC_PATH.relative_to(ROOT)}")
    print(f"- Source-pack completion: {source_pack_completion_summary(validation)}")
    if validation.issues:
        for issue in validation.issues:
            print(f"- {issue}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
