from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook

from generate_slic_workbook import WORKBOOK_PATH, build_city_universe, unique_countries


ROOT = Path(__file__).resolve().parent.parent
VERIFIED_SOURCE_DIR = ROOT / "data" / "verified_sources"
PROVIDERS_PATH = VERIFIED_SOURCE_DIR / "providers.csv"
COUNTRY_CONTEXT_PATH = VERIFIED_SOURCE_DIR / "country_context.csv"
CITY_INPUTS_PATH = VERIFIED_SOURCE_DIR / "city_inputs.csv"
COUNTRY_CONTEXT_FIELDS = [
    "country",
    "field",
    "label",
    "pillar",
    "required_for_ranking",
    "value",
    "provider_id",
    "source_url",
    "source_title",
    "source_date",
    "reference_period",
    "source_scope",
    "proxy_status",
    "notes",
]
CITY_INPUT_FIELDS = [
    "city_id",
    "display_name",
    "country",
    "cohort",
    "field",
    "label",
    "pillar",
    "required_for_ranking",
    "value",
    "provider_id",
    "source_url",
    "source_title",
    "source_date",
    "reference_period",
    "source_scope",
    "proxy_status",
    "notes",
]

PILLAR_WEIGHTS = {
    "pressure": 25,
    "viability": 22,
    "capability": 18,
    "community": 15,
    "creative": 20,
}
PILLAR_ORDER = ("pressure", "viability", "capability", "community", "creative")
TIER_ORDER = {"Tier 1": 1, "Tier 2": 2, "Tier 3": 3, "Tier 4": 4, "Tier 5": 5}
PUBLIC_PROVIDER_TIERS = {"Tier 1", "Tier 2", "Tier 3"}
SOURCE_SCOPE_VALUES = ("city", "metro", "regional", "national", "international")
PROXY_STATUS_VALUES = ("direct", "approved_proxy")
FRESHNESS_LIMITS_MONTHS = {
    "fast": 24,
    "medium": 36,
    "structural": 60,
}
PROVIDER_FIELDS = [
    "provider_id",
    "name",
    "tier",
    "scope",
    "allowed_host",
    "allow_any_https",
    "reference_only",
    "notes",
]

STARTER_PROVIDERS = [
    {
        "provider_id": "city_official_portal",
        "name": "City or metro official portal",
        "tier": "Tier 1",
        "scope": "city_or_metro",
        "allowed_host": "",
        "allow_any_https": "true",
        "reference_only": "false",
        "notes": "Use for municipal statistical offices, city open-data portals, transit authorities, or water and utility agencies.",
    },
    {
        "provider_id": "subnational_official_portal",
        "name": "State or provincial official portal",
        "tier": "Tier 2",
        "scope": "subnational",
        "allowed_host": "",
        "allow_any_https": "true",
        "reference_only": "false",
        "notes": "Use for provincial, prefectural, or metropolitan statistical offices and observatories.",
    },
    {
        "provider_id": "national_statistical_office",
        "name": "National statistical office",
        "tier": "Tier 3",
        "scope": "national",
        "allowed_host": "",
        "allow_any_https": "true",
        "reference_only": "false",
        "notes": "Use for official national statistics portals when a city metric is only available nationally or by subnational breakdown.",
    },
    {
        "provider_id": "world_bank",
        "name": "World Bank Data",
        "tier": "Tier 3",
        "scope": "international",
        "allowed_host": "worldbank.org",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Recommended for GDP per capita PPP, growth, inequality context, and related official international baselines.",
    },
    {
        "provider_id": "istat",
        "name": "ISTAT",
        "tier": "Tier 3",
        "scope": "national",
        "allowed_host": "istat.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Italian national statistical institute. Use for labour, prices, safety, education, and subnational social indicators.",
    },
    {
        "provider_id": "banca_d_italia",
        "name": "Banca d'Italia",
        "tier": "Tier 3",
        "scope": "national",
        "allowed_host": "bancaditalia.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Bank of Italy source for credit, household finance, and territorial banking indicators.",
    },
    {
        "provider_id": "agenzia_entrate_omi",
        "name": "Agenzia delle Entrate OMI",
        "tier": "Tier 3",
        "scope": "national",
        "allowed_host": "agenziaentrate.gov.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Property Market Observatory for housing and rent reference values in Italian cities.",
    },
    {
        "provider_id": "comune_milano_open_data",
        "name": "Comune di Milano Open Data",
        "tier": "Tier 1",
        "scope": "city_or_metro",
        "allowed_host": "dati.comune.milano.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Comune di Milano open-data portal for city statistics, mobility, services, business activity, and civic indicators.",
    },
    {
        "provider_id": "atm_milano",
        "name": "ATM Milano",
        "tier": "Tier 1",
        "scope": "city_or_metro",
        "allowed_host": "atm.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Milan public-transport operator for fares, service coverage, and network information.",
    },
    {
        "provider_id": "arpa_lombardia",
        "name": "ARPA Lombardia",
        "tier": "Tier 2",
        "scope": "subnational",
        "allowed_host": "arpalombardia.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Lombardy regional environmental protection agency for air quality and environmental monitoring.",
    },
    {
        "provider_id": "mm_milano",
        "name": "MM S.p.A.",
        "tier": "Tier 2",
        "scope": "utility_operator",
        "allowed_host": "mmspa.eu",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Milan utility operator for water-service and network performance information.",
    },
    {
        "provider_id": "comune_bologna",
        "name": "Comune di Bologna",
        "tier": "Tier 1",
        "scope": "city_or_metro",
        "allowed_host": "comune.bologna.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Comune di Bologna portal for city programmes, civic services, culture, and local statistics.",
    },
    {
        "provider_id": "comune_bologna_open_data",
        "name": "Comune di Bologna Open Data",
        "tier": "Tier 1",
        "scope": "city_or_metro",
        "allowed_host": "opendata.comune.bologna.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Comune di Bologna open-data portal for mobility, demographics, business, neighbourhood, and service indicators.",
    },
    {
        "provider_id": "tper_bologna",
        "name": "TPER Bologna",
        "tier": "Tier 1",
        "scope": "city_or_metro",
        "allowed_host": "tper.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Bologna-area transport operator for fares, service information, and local transit network details.",
    },
    {
        "provider_id": "arpae_emilia_romagna",
        "name": "ARPAE Emilia-Romagna",
        "tier": "Tier 2",
        "scope": "subnational",
        "allowed_host": "arpae.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Official Emilia-Romagna environmental agency for air, water, and local environmental quality monitoring.",
    },
    {
        "provider_id": "gruppo_hera",
        "name": "Gruppo Hera",
        "tier": "Tier 2",
        "scope": "utility_operator",
        "allowed_host": "gruppohera.it",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Primary regional utility operator for Bologna-area water, waste, and utility service data.",
    },
    {
        "provider_id": "who_gho",
        "name": "WHO Global Health Observatory",
        "tier": "Tier 3",
        "scope": "international",
        "allowed_host": "who.int",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Recommended for health and mortality baselines when city-level official releases are unavailable.",
    },
    {
        "provider_id": "ilo_ilostat",
        "name": "ILO ILOSTAT",
        "tier": "Tier 3",
        "scope": "international",
        "allowed_host": "ilo.org",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Recommended for labour, work-time, and employment pressure indicators.",
    },
    {
        "provider_id": "who_unicef_jmp",
        "name": "WHO/UNICEF JMP",
        "tier": "Tier 3",
        "scope": "international",
        "allowed_host": "washdata.org",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Recommended for water, sanitation, and utility service baselines.",
    },
    {
        "provider_id": "unesco_uis",
        "name": "UNESCO UIS",
        "tier": "Tier 3",
        "scope": "international",
        "allowed_host": "uis.unesco.org",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Recommended for education and human-capability indicators.",
    },
    {
        "provider_id": "oecd_data",
        "name": "OECD Data",
        "tier": "Tier 3",
        "scope": "international",
        "allowed_host": "oecd.org",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Recommended for housing, health, education, and competitive-context baselines.",
    },
    {
        "provider_id": "wipo_statistics",
        "name": "WIPO IP Statistics",
        "tier": "Tier 3",
        "scope": "international",
        "allowed_host": "wipo.int",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Recommended for innovation and research intensity proxies.",
    },
    {
        "provider_id": "openaq",
        "name": "OpenAQ",
        "tier": "Tier 4",
        "scope": "secondary",
        "allowed_host": "openaq.org",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Audited secondary layer for air-quality coverage when official city feeds are thin.",
    },
    {
        "provider_id": "mlab",
        "name": "Measurement Lab",
        "tier": "Tier 4",
        "scope": "secondary",
        "allowed_host": "measurementlab.net",
        "allow_any_https": "false",
        "reference_only": "false",
        "notes": "Audited secondary layer for digital-performance context.",
    },
    {
        "provider_id": "economist_world_in_numbers",
        "name": "The World in Numbers | The Economist",
        "tier": "Tier 4",
        "scope": "secondary_editorial_reference",
        "allowed_host": "economist.com",
        "allow_any_https": "false",
        "reference_only": "true",
        "notes": "Internal analyst reference only. Do not use this provider directly in publishable Country_Context or City_Inputs scoring rows.",
    },
]

COUNTRY_FIELD_SPECS = [
    {
        "field": "gdp_per_capita_ppp",
        "label": "GDP per capita (PPP)",
        "pillar": "creative",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "gdp_growth",
        "label": "GDP growth",
        "pillar": "creative",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "gini_coefficient",
        "label": "Gini coefficient",
        "pillar": "capability",
        "required_for_ranking": "yes",
        "freshness_class": "structural",
    },
    {
        "field": "ppp_private_consumption",
        "label": "PPP private consumption factor",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "tax_rate_assumption",
        "label": "Tax rate assumption",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "household_debt_proxy",
        "label": "Household debt proxy",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
]

CITY_FIELD_SPECS = [
    {
        "field": "gross_income",
        "label": "Gross income",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "rent",
        "label": "Rent",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "utilities",
        "label": "Utilities",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "transit_cost",
        "label": "Transit cost",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "internet_cost",
        "label": "Internet cost",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "food_cost",
        "label": "Food cost",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "housing_burden_raw",
        "label": "Housing burden raw",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "household_debt_burden_raw",
        "label": "Household debt burden raw",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "working_time_pressure_raw",
        "label": "Working time pressure raw",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "suicide_mental_strain_raw",
        "label": "Suicide or mental strain raw",
        "pillar": "pressure",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "personal_safety_raw",
        "label": "Personal safety raw",
        "pillar": "viability",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "transit_access_commute_raw",
        "label": "Transit access and commute raw",
        "pillar": "viability",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "clean_air_raw",
        "label": "Clean air raw",
        "pillar": "viability",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "water_sanitation_utility_raw",
        "label": "Water, sanitation, and utility raw",
        "pillar": "viability",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "digital_infrastructure_raw",
        "label": "Digital infrastructure raw",
        "pillar": "viability",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "climate_sunlight_livability_raw",
        "label": "Climate and sunlight livability raw",
        "pillar": "viability",
        "required_for_ranking": "yes",
        "freshness_class": "structural",
    },
    {
        "field": "healthcare_quality_raw",
        "label": "Healthcare quality raw",
        "pillar": "capability",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "education_quality_raw",
        "label": "Education quality raw",
        "pillar": "capability",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "equal_opportunity_raw",
        "label": "Equal opportunity raw",
        "pillar": "capability",
        "required_for_ranking": "yes",
        "freshness_class": "structural",
    },
    {
        "field": "hospitality_belonging_raw",
        "label": "Hospitality and belonging raw",
        "pillar": "community",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "tolerance_pluralism_raw",
        "label": "Tolerance and pluralism raw",
        "pillar": "community",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "cultural_public_life_raw",
        "label": "Cultural public life raw",
        "pillar": "community",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "birth_rate_optimism_raw",
        "label": "Birth rate and societal optimism raw",
        "pillar": "community",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "entrepreneurial_dynamism_raw",
        "label": "Entrepreneurial dynamism raw",
        "pillar": "creative",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "innovation_research_intensity_raw",
        "label": "Innovation and research raw",
        "pillar": "creative",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "investment_signal_raw",
        "label": "Investment signal raw",
        "pillar": "creative",
        "required_for_ranking": "yes",
        "freshness_class": "fast",
    },
    {
        "field": "administrative_investment_friction_raw",
        "label": "Administrative investment friction raw",
        "pillar": "creative",
        "required_for_ranking": "yes",
        "freshness_class": "medium",
    },
    {
        "field": "visitor_flow_context_raw",
        "label": "Visitor flow context raw",
        "pillar": "creative",
        "required_for_ranking": "no",
        "freshness_class": "fast",
    },
]

CITY_SOURCE_COLUMNS = [
    "pressure_source_url",
    "pressure_source_tier",
    "pressure_source_date",
    "pressure_notes",
    "viability_source_url",
    "viability_source_tier",
    "viability_source_date",
    "viability_notes",
    "capability_source_url",
    "capability_source_tier",
    "capability_source_date",
    "capability_notes",
    "community_source_url",
    "community_source_tier",
    "community_source_date",
    "community_notes",
    "creative_source_url",
    "creative_source_tier",
    "creative_source_date",
    "creative_notes",
]

COUNTRY_FIELD_LOOKUP = {spec["field"]: spec for spec in COUNTRY_FIELD_SPECS}
CITY_FIELD_LOOKUP = {spec["field"]: spec for spec in CITY_FIELD_SPECS}
COUNTRY_CONTEXT_FIELD_MAP = {
    "gdp_per_capita_ppp": "gdp_per_capita_ppp_context",
    "gdp_growth": "gdp_growth_context",
    "gini_coefficient": "gini_coefficient_context",
    "ppp_private_consumption": "ppp_private_consumption_context",
    "tax_rate_assumption": "tax_rate_context",
    "household_debt_proxy": "household_debt_proxy_context",
}
CITY_FIELDS_BY_PILLAR = {
    pillar: [spec["field"] for spec in CITY_FIELD_SPECS if spec["pillar"] == pillar]
    for pillar in PILLAR_ORDER
}
COUNTRY_FIELDS_BY_PILLAR = {
    pillar: [spec["field"] for spec in COUNTRY_FIELD_SPECS if spec["pillar"] == pillar]
    for pillar in PILLAR_ORDER
}
CITY_FIELD_ORDER = [spec["field"] for spec in CITY_FIELD_SPECS]
COUNTRY_FIELD_ORDER = [spec["field"] for spec in COUNTRY_FIELD_SPECS]

METRIC_SPECS: dict[str, dict[str, object]] = {
    "pressure_disposable_income_ppp": {"input_key": "di_ppp_raw", "weight": 9, "type": "direct"},
    "pressure_housing_burden": {"input_key": "housing_burden_raw", "weight": 5, "type": "direct"},
    "pressure_household_debt_burden": {"input_key": "household_debt_effective_raw", "weight": 4, "type": "direct"},
    "pressure_working_time_pressure": {"input_key": "working_time_pressure_raw", "weight": 4, "type": "direct"},
    "pressure_suicide_mental_strain": {"input_key": "suicide_mental_strain_raw", "weight": 5, "type": "direct"},
    "viability_personal_safety": {"input_key": "personal_safety_raw", "weight": 5, "type": "direct"},
    "viability_transit_access_commute": {"input_key": "transit_access_commute_raw", "weight": 5, "type": "direct"},
    "viability_clean_air": {"input_key": "clean_air_raw", "weight": 4, "type": "direct"},
    "viability_water_sanitation_utility": {"input_key": "water_sanitation_utility_raw", "weight": 4, "type": "direct"},
    "viability_digital_infrastructure": {"input_key": "digital_infrastructure_raw", "weight": 4, "type": "direct"},
    "viability_climate_sunlight_livability": {"input_key": "climate_sunlight_livability_raw", "weight": 5, "type": "direct"},
    "capability_healthcare_quality": {"input_key": "healthcare_quality_raw", "weight": 8, "type": "direct"},
    "capability_education_quality": {"input_key": "education_quality_raw", "weight": 6, "type": "direct"},
    "capability_equal_opportunity_distributional_fairness": {
        "weight": 4,
        "type": "composite",
        "components": [
            ("equal_opportunity_raw", 0.7),
            ("gini_coefficient_context", 0.3),
        ],
    },
    "community_hospitality_belonging": {"input_key": "hospitality_belonging_raw", "weight": 5, "type": "direct"},
    "community_tolerance_pluralism": {"input_key": "tolerance_pluralism_raw", "weight": 5, "type": "direct"},
    "community_cultural_historic_public_life_vitality": {
        "input_key": "cultural_public_life_raw",
        "weight": 5,
        "type": "direct",
    },
    "community_birth_rate_optimism": {"input_key": "birth_rate_optimism_raw", "weight": 4, "type": "direct"},
    "creative_entrepreneurial_dynamism": {"input_key": "entrepreneurial_dynamism_raw", "weight": 6, "type": "direct"},
    "creative_innovation_research_intensity": {
        "input_key": "innovation_research_intensity_raw",
        "weight": 5,
        "type": "direct",
    },
    "creative_economic_vitality_productive_context": {
        "weight": 5,
        "type": "composite",
        "components": [
            ("investment_signal_raw", 0.5),
            ("gdp_per_capita_ppp_context", 0.3),
            ("gdp_growth_context", 0.2),
        ],
    },
    "creative_administrative_investment_friction": {
        "input_key": "administrative_investment_friction_raw",
        "weight": 4,
        "type": "direct",
    },
}

PILLAR_METRICS = {
    "pressure": [
        ("pressure_disposable_income_ppp", 9),
        ("pressure_housing_burden", 5),
        ("pressure_household_debt_burden", 4),
        ("pressure_working_time_pressure", 4),
        ("pressure_suicide_mental_strain", 5),
    ],
    "viability": [
        ("viability_personal_safety", 5),
        ("viability_transit_access_commute", 5),
        ("viability_clean_air", 4),
        ("viability_water_sanitation_utility", 4),
        ("viability_digital_infrastructure", 4),
        ("viability_climate_sunlight_livability", 5),
    ],
    "capability": [
        ("capability_healthcare_quality", 8),
        ("capability_education_quality", 6),
        ("capability_equal_opportunity_distributional_fairness", 4),
    ],
    "community": [
        ("community_hospitality_belonging", 5),
        ("community_tolerance_pluralism", 5),
        ("community_cultural_historic_public_life_vitality", 5),
        ("community_birth_rate_optimism", 4),
    ],
    "creative": [
        ("creative_entrepreneurial_dynamism", 6),
        ("creative_innovation_research_intensity", 5),
        ("creative_economic_vitality_productive_context", 5),
        ("creative_administrative_investment_friction", 4),
    ],
}

SCORE_INPUTS = {
    "di_ppp_raw": {"source_column": "di_ppp_raw", "directionality": "positive"},
    "housing_burden_raw": {"source_column": "housing_burden_raw", "directionality": "negative"},
    "household_debt_effective_raw": {"source_column": "household_debt_effective_raw", "directionality": "negative"},
    "working_time_pressure_raw": {"source_column": "working_time_pressure_raw", "directionality": "negative"},
    "suicide_mental_strain_raw": {"source_column": "suicide_mental_strain_raw", "directionality": "negative"},
    "personal_safety_raw": {"source_column": "personal_safety_raw", "directionality": "negative"},
    "transit_access_commute_raw": {"source_column": "transit_access_commute_raw", "directionality": "positive"},
    "clean_air_raw": {"source_column": "clean_air_raw", "directionality": "negative"},
    "water_sanitation_utility_raw": {"source_column": "water_sanitation_utility_raw", "directionality": "positive"},
    "digital_infrastructure_raw": {"source_column": "digital_infrastructure_raw", "directionality": "positive"},
    "climate_sunlight_livability_raw": {"source_column": "climate_sunlight_livability_raw", "directionality": "positive"},
    "healthcare_quality_raw": {"source_column": "healthcare_quality_raw", "directionality": "positive"},
    "education_quality_raw": {"source_column": "education_quality_raw", "directionality": "positive"},
    "equal_opportunity_raw": {"source_column": "equal_opportunity_raw", "directionality": "positive"},
    "gini_coefficient_context": {"source_column": "gini_coefficient_context", "directionality": "negative"},
    "hospitality_belonging_raw": {"source_column": "hospitality_belonging_raw", "directionality": "positive"},
    "tolerance_pluralism_raw": {"source_column": "tolerance_pluralism_raw", "directionality": "positive"},
    "cultural_public_life_raw": {"source_column": "cultural_public_life_raw", "directionality": "positive"},
    "entrepreneurial_dynamism_raw": {"source_column": "entrepreneurial_dynamism_raw", "directionality": "positive"},
    "innovation_research_intensity_raw": {"source_column": "innovation_research_intensity_raw", "directionality": "positive"},
    "investment_signal_raw": {"source_column": "investment_signal_raw", "directionality": "positive"},
    "gdp_per_capita_ppp_context": {"source_column": "gdp_per_capita_ppp_context", "directionality": "positive"},
    "gdp_growth_context": {"source_column": "gdp_growth_context", "directionality": "positive"},
    "administrative_investment_friction_raw": {
        "source_column": "administrative_investment_friction_raw",
        "directionality": "negative",
    },
    "birth_rate_optimism_raw": {"source_column": "birth_rate_optimism_raw", "directionality": "positive"},
}


@dataclass(frozen=True)
class Provider:
    provider_id: str
    name: str
    tier: str
    scope: str
    allowed_host: str
    allow_any_https: bool
    reference_only: bool
    notes: str


@dataclass(frozen=True)
class SourceEntry:
    field: str
    value: float
    provider_id: str
    tier: str
    source_url: str
    source_title: str
    source_date: str
    reference_period: str
    source_scope: str
    proxy_status: str
    notes: str


@dataclass(frozen=True)
class IntegritySummary:
    key: str
    label: str
    count: int
    samples: tuple[str, ...]


@dataclass
class SourcePackValidation:
    providers: dict[str, Provider]
    country_values: dict[str, dict[str, SourceEntry]]
    city_values: dict[str, dict[str, SourceEntry]]
    issues: list[str]
    integrity_summaries: list[IntegritySummary]
    stats: dict[str, int]


def bool_from_text(value: str) -> bool:
    return value.strip().lower() in {"1", "true", "yes", "y"}


class IssueRecorder:
    def __init__(self) -> None:
        self._counts: dict[str, int] = {}
        self._labels: dict[str, str] = {}
        self._samples: dict[str, list[str]] = {}

    def add(self, key: str, label: str, sample: str) -> None:
        self._counts[key] = self._counts.get(key, 0) + 1
        self._labels[key] = label
        bucket = self._samples.setdefault(key, [])
        if len(bucket) < 4 and sample not in bucket:
            bucket.append(sample)

    def summaries(self) -> list[IntegritySummary]:
        return [
            IntegritySummary(
                key=key,
                label=self._labels[key],
                count=self._counts[key],
                samples=tuple(self._samples.get(key, [])),
            )
            for key in sorted(self._counts)
        ]


def summarize_integrity_issues(summaries: list[IntegritySummary]) -> list[str]:
    issues: list[str] = []
    for summary in summaries:
        line = f"{summary.label}: {summary.count}"
        if summary.samples:
            line += f". Examples: {'; '.join(summary.samples)}"
        issues.append(line)
    return issues


def months_old(source_date: date, today: date) -> int:
    months = (today.year - source_date.year) * 12 + (today.month - source_date.month)
    if today.day < source_date.day:
        months -= 1
    return months


def provider_scope_to_source_scope(provider: Provider, row_kind: str) -> str:
    scope = provider.scope.strip().lower()
    if scope == "city_or_metro":
        return "city"
    if scope == "subnational":
        return "regional"
    if scope == "utility_operator":
        return "city" if row_kind == "city" else "regional"
    if scope == "national":
        return "national"
    if scope == "national_international":
        return "international"
    if scope == "international":
        return "international"
    return ""


def default_reference_period(source_date: str) -> str:
    if not source_date:
        return ""
    return f"as_of {source_date}"


def default_proxy_status(
    row_kind: str,
    field: str,
    source_scope: str,
    provider: Provider | None,
) -> str:
    if field.endswith("_proxy") or field.endswith("_assumption") or "proxy" in field or "assumption" in field:
        return "approved_proxy"
    if provider and provider.tier in {"Tier 4", "Tier 5"}:
        return "approved_proxy"
    if row_kind == "city" and source_scope not in {"city", "metro"}:
        return "approved_proxy"
    return "direct"


def normalize_row_metadata(
    row: dict[str, str],
    fieldnames: list[str],
    provider_map: dict[str, Provider],
    row_kind: str,
) -> dict[str, str]:
    normalized = {field: row.get(field, "").strip() for field in fieldnames}
    provider_id = normalized.get("provider_id", "")
    provider = provider_map.get(provider_id)
    if provider is not None:
        if "reference_period" in normalized and not normalized["reference_period"] and normalized.get("source_date", ""):
            normalized["reference_period"] = default_reference_period(normalized["source_date"])
        if "source_scope" in normalized and not normalized["source_scope"]:
            normalized["source_scope"] = provider_scope_to_source_scope(provider, row_kind)
        if "proxy_status" in normalized and not normalized["proxy_status"]:
            normalized["proxy_status"] = default_proxy_status(
                row_kind=row_kind,
                field=normalized.get("field", ""),
                source_scope=normalized.get("source_scope", ""),
                provider=provider,
            )
    return normalized


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def merge_expected_rows(
    existing_rows: list[dict[str, str]],
    expected_rows: list[dict[str, str]],
    key_fields: tuple[str, ...],
    fieldnames: list[str],
) -> list[dict[str, str]]:
    expected_by_key = {
        tuple(expected_row[field] for field in key_fields): expected_row
        for expected_row in expected_rows
    }
    consumed_keys: set[tuple[str, ...]] = set()
    extras: list[dict[str, str]] = []

    for row in existing_rows:
        key = tuple(row.get(field, "").strip() for field in key_fields)
        if key not in expected_by_key or key in consumed_keys:
            extras.append({field: row.get(field, "").strip() for field in fieldnames})
            continue
        consumed_keys.add(key)

    existing_by_key = {
        tuple(row.get(field, "").strip() for field in key_fields): row
        for row in existing_rows
        if tuple(row.get(field, "").strip() for field in key_fields) in expected_by_key
    }

    merged_rows: list[dict[str, str]] = []
    for expected_row in expected_rows:
        key = tuple(expected_row[field] for field in key_fields)
        existing_row = existing_by_key.get(key)
        merged_row = {field: expected_row.get(field, "") for field in fieldnames}
        if existing_row is not None:
            for field in fieldnames:
                existing_value = existing_row.get(field, "").strip()
                if existing_value != "":
                    merged_row[field] = existing_value
        merged_rows.append(merged_row)

    merged_rows.extend(extras)
    return merged_rows


def prepare_verified_source_pack(force: bool = False) -> dict[str, Path]:
    VERIFIED_SOURCE_DIR.mkdir(parents=True, exist_ok=True)

    if force or not PROVIDERS_PATH.exists():
        write_csv(PROVIDERS_PATH, STARTER_PROVIDERS, PROVIDER_FIELDS)
    else:
        existing_rows = read_csv_rows(PROVIDERS_PATH)
        merged_rows: list[dict[str, str]] = []
        existing_by_id = {row.get("provider_id", "").strip(): row for row in existing_rows}
        written_ids: set[str] = set()

        for row in existing_rows:
            provider_id = row.get("provider_id", "").strip()
            if not provider_id:
                continue
            starter = next((item for item in STARTER_PROVIDERS if item["provider_id"] == provider_id), None)
            merged = {field: row.get(field, "").strip() for field in PROVIDER_FIELDS}
            if starter:
                for field in PROVIDER_FIELDS:
                    if merged[field] == "":
                        merged[field] = starter.get(field, "")
            merged_rows.append(merged)
            written_ids.add(provider_id)

        for starter in STARTER_PROVIDERS:
            if starter["provider_id"] in written_ids:
                continue
            merged_rows.append({field: starter.get(field, "") for field in PROVIDER_FIELDS})

        write_csv(PROVIDERS_PATH, merged_rows, PROVIDER_FIELDS)

    provider_map, _ = load_provider_registry()

    city_universe = build_city_universe()
    countries = unique_countries(city_universe)

    country_rows = [
        {
            "country": country,
            "field": spec["field"],
            "label": spec["label"],
            "pillar": spec["pillar"],
            "required_for_ranking": spec["required_for_ranking"],
            "value": "",
            "provider_id": "",
            "source_url": "",
            "source_title": "",
            "source_date": "",
            "reference_period": "",
            "source_scope": "",
            "proxy_status": "",
            "notes": "",
        }
        for country in countries
        for spec in COUNTRY_FIELD_SPECS
    ]
    if force or not COUNTRY_CONTEXT_PATH.exists():
        write_csv(
            COUNTRY_CONTEXT_PATH,
            [
                normalize_row_metadata(row, COUNTRY_CONTEXT_FIELDS, provider_map, row_kind="country")
                for row in country_rows
            ],
            COUNTRY_CONTEXT_FIELDS,
        )
    else:
        write_csv(
            COUNTRY_CONTEXT_PATH,
            [
                normalize_row_metadata(row, COUNTRY_CONTEXT_FIELDS, provider_map, row_kind="country")
                for row in merge_expected_rows(
                    read_csv_rows(COUNTRY_CONTEXT_PATH),
                    country_rows,
                    ("country", "field"),
                    COUNTRY_CONTEXT_FIELDS,
                )
            ],
            COUNTRY_CONTEXT_FIELDS,
        )

    city_rows = [
        {
            "city_id": city["city_id"],
            "display_name": city["display_name"],
            "country": city["country"],
            "cohort": city["cohort"],
            "field": spec["field"],
            "label": spec["label"],
            "pillar": spec["pillar"],
            "required_for_ranking": spec["required_for_ranking"],
            "value": "",
            "provider_id": "",
            "source_url": "",
            "source_title": "",
            "source_date": "",
            "reference_period": "",
            "source_scope": "",
            "proxy_status": "",
            "notes": "",
        }
        for city in city_universe
        for spec in CITY_FIELD_SPECS
    ]
    if force or not CITY_INPUTS_PATH.exists():
        write_csv(
            CITY_INPUTS_PATH,
            [
                normalize_row_metadata(row, CITY_INPUT_FIELDS, provider_map, row_kind="city")
                for row in city_rows
            ],
            CITY_INPUT_FIELDS,
        )
    else:
        write_csv(
            CITY_INPUTS_PATH,
            [
                normalize_row_metadata(row, CITY_INPUT_FIELDS, provider_map, row_kind="city")
                for row in merge_expected_rows(
                    read_csv_rows(CITY_INPUTS_PATH),
                    city_rows,
                    ("city_id", "field"),
                    CITY_INPUT_FIELDS,
                )
            ],
            CITY_INPUT_FIELDS,
        )

    return {
        "providers": PROVIDERS_PATH,
        "country_context": COUNTRY_CONTEXT_PATH,
        "city_inputs": CITY_INPUTS_PATH,
    }


def host_matches(url: str, allowed_host: str) -> bool:
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower()
    allowed = allowed_host.lower()
    return host == allowed or host.endswith(f".{allowed}")


def load_provider_registry() -> tuple[dict[str, Provider], list[str]]:
    issues: list[str] = []
    providers: dict[str, Provider] = {}

    if not PROVIDERS_PATH.exists():
        return {}, [f"Missing provider registry: {PROVIDERS_PATH.relative_to(ROOT)}"]

    for row_index, row in enumerate(read_csv_rows(PROVIDERS_PATH), start=2):
        provider_id = row.get("provider_id", "").strip()
        if not provider_id:
            issues.append(f"providers.csv row {row_index} is missing provider_id.")
            continue
        if provider_id in providers:
            issues.append(f"providers.csv row {row_index} duplicates provider_id '{provider_id}'.")
            continue

        tier = row.get("tier", "").strip()
        if tier not in TIER_ORDER:
            issues.append(f"providers.csv row {row_index} has invalid tier '{tier}'.")
            continue

        providers[provider_id] = Provider(
            provider_id=provider_id,
            name=row.get("name", "").strip(),
            tier=tier,
            scope=row.get("scope", "").strip(),
            allowed_host=row.get("allowed_host", "").strip(),
            allow_any_https=bool_from_text(row.get("allow_any_https", "")),
            reference_only=bool_from_text(row.get("reference_only", "false")),
            notes=row.get("notes", "").strip(),
        )

    return providers, issues


def validate_entry(
    row: dict[str, str],
    row_label: str,
    provider_map: dict[str, Provider],
    field_spec: dict[str, str],
    row_kind: str,
    issue_recorder: IssueRecorder,
    today: date,
) -> tuple[SourceEntry | None, list[str]]:
    issues: list[str] = []
    value_text = row.get("value", "").strip()
    provider_id = row.get("provider_id", "").strip()
    source_url = row.get("source_url", "").strip()
    source_title = row.get("source_title", "").strip()
    source_date = row.get("source_date", "").strip()
    reference_period = row.get("reference_period", "").strip()
    source_scope = row.get("source_scope", "").strip()
    proxy_status = row.get("proxy_status", "").strip()
    notes = row.get("notes", "").strip()

    if not any(
        (
            value_text,
            provider_id,
            source_url,
            source_title,
            source_date,
            reference_period,
            source_scope,
            proxy_status,
            notes,
        )
    ):
        return None, issues

    if not value_text:
        issue_recorder.add(
            "missing_numeric_value",
            "Rows with source metadata but no numeric value",
            row_label,
        )
    if not provider_id:
        issue_recorder.add("missing_provider_id", "Rows missing provider_id", row_label)
    if not source_url:
        issue_recorder.add("missing_source_url", "Rows missing source_url", row_label)
    if not source_title:
        issue_recorder.add("missing_source_title", "Rows missing source_title", row_label)
    if not source_date:
        issue_recorder.add("missing_source_date", "Rows missing source_date", row_label)
    if not reference_period:
        issue_recorder.add("missing_reference_period", "Rows missing reference_period", row_label)
    if not source_scope:
        issue_recorder.add("missing_source_scope", "Rows missing source_scope", row_label)
    if not proxy_status:
        issue_recorder.add("missing_proxy_status", "Rows missing proxy_status", row_label)

    if not value_text or not provider_id or not source_url or not source_title or not source_date or not reference_period or not source_scope or not proxy_status:
        return None, issues

    try:
        value = float(value_text)
    except ValueError:
        issue_recorder.add("non_numeric_value", "Rows with non-numeric values", f"{row_label} -> {value_text}")
        return None, issues

    provider = provider_map.get(provider_id)
    if provider is None:
        issue_recorder.add(
            "unknown_provider",
            "Rows referencing unknown provider_id values",
            f"{row_label} -> {provider_id}",
        )
        return None, issues
    if provider.reference_only:
        issue_recorder.add(
            "reference_only_provider",
            "Rows using reference-only providers in scoring inputs",
            f"{row_label} -> {provider_id}",
        )
        return None, issues

    required_for_ranking = field_spec["required_for_ranking"] == "yes"
    if required_for_ranking and provider.tier not in PUBLIC_PROVIDER_TIERS:
        issue_recorder.add(
            "non_publishable_tier",
            "Required scoring rows using Tier 4 or Tier 5 providers",
            f"{row_label} -> {provider_id} ({provider.tier})",
        )
        return None, issues

    parsed = urlparse(source_url)
    if parsed.scheme != "https" or not parsed.netloc:
        issue_recorder.add("invalid_https_url", "Rows with invalid https source_url values", row_label)
        return None, issues

    if provider.allowed_host and not host_matches(source_url, provider.allowed_host):
        issue_recorder.add(
            "provider_host_mismatch",
            "Rows whose source_url host does not match provider policy",
            f"{row_label} -> {parsed.hostname or ''}",
        )
        return None, issues

    if not provider.allow_any_https and not provider.allowed_host:
        issues.append(f"Provider '{provider_id}' does not allow arbitrary hosts and is missing allowed_host.")
        return None, issues

    try:
        parsed_source_date = date.fromisoformat(source_date)
    except ValueError:
        issue_recorder.add(
            "invalid_source_date",
            "Rows with invalid source_date values",
            f"{row_label} -> {source_date}",
        )
        return None, issues

    if source_scope not in SOURCE_SCOPE_VALUES:
        issue_recorder.add(
            "invalid_source_scope",
            "Rows with invalid source_scope values",
            f"{row_label} -> {source_scope}",
        )
        return None, issues
    if proxy_status not in PROXY_STATUS_VALUES:
        issue_recorder.add(
            "invalid_proxy_status",
            "Rows with invalid proxy_status values",
            f"{row_label} -> {proxy_status}",
        )
        return None, issues

    field_name = row["field"].strip()
    if row_kind == "city" and source_scope not in {"city", "metro"} and proxy_status != "approved_proxy":
        issue_recorder.add(
            "city_scope_requires_proxy",
            "City rows using broader-than-city scope without approved_proxy",
            row_label,
        )
        return None, issues
    if row_kind == "country" and ("proxy" in field_name or "assumption" in field_name) and proxy_status != "approved_proxy":
        issue_recorder.add(
            "country_proxy_requires_approved_proxy",
            "Country proxy rows not marked approved_proxy",
            row_label,
        )
        return None, issues

    freshness_class = str(field_spec["freshness_class"])
    limit_months = FRESHNESS_LIMITS_MONTHS[freshness_class]
    age_months = months_old(parsed_source_date, today)
    if age_months > limit_months:
        issue_recorder.add(
            f"stale_{freshness_class}",
            f"Rows older than the {limit_months}-month freshness limit",
            f"{row_label} -> {source_date}",
        )
        return None, issues
    if freshness_class == "structural" and age_months > 36 and proxy_status != "approved_proxy":
        issue_recorder.add(
            "structural_stale_without_proxy",
            "Structural rows older than 36 months without approved_proxy",
            f"{row_label} -> {source_date}",
        )
        return None, issues

    return (
        SourceEntry(
            field=field_name,
            value=value,
            provider_id=provider_id,
            tier=provider.tier,
            source_url=source_url,
            source_title=source_title,
            source_date=source_date,
            reference_period=reference_period,
            source_scope=source_scope,
            proxy_status=proxy_status,
            notes=notes,
        ),
        issues,
    )


def validate_source_pack() -> SourcePackValidation:
    prepare_verified_source_pack(force=False)
    provider_map, issues = load_provider_registry()
    issue_recorder = IssueRecorder()
    today = date.today()

    city_universe = build_city_universe()
    countries = unique_countries(city_universe)
    city_ids = [city["city_id"] for city in city_universe]

    country_rows = read_csv_rows(COUNTRY_CONTEXT_PATH)
    city_rows = read_csv_rows(CITY_INPUTS_PATH)

    expected_country_keys = {(country, field) for country in countries for field in COUNTRY_FIELD_ORDER}
    expected_city_keys = {(city_id, field) for city_id in city_ids for field in CITY_FIELD_ORDER}

    observed_country_keys: set[tuple[str, str]] = set()
    observed_city_keys: set[tuple[str, str]] = set()
    country_values: dict[str, dict[str, SourceEntry]] = {}
    city_values: dict[str, dict[str, SourceEntry]] = {}
    filled_country_rows = 0
    filled_city_rows = 0

    for row_index, row in enumerate(country_rows, start=2):
        country = row.get("country", "").strip()
        field = row.get("field", "").strip()
        row_label = f"country_context.csv row {row_index} ({country or 'unknown'} / {field or 'unknown'})"
        if country not in countries:
            issues.append(f"{row_label} references unknown country '{country}'.")
            continue
        if field not in COUNTRY_FIELD_LOOKUP:
            issues.append(f"{row_label} references unknown field '{field}'.")
            continue

        key = (country, field)
        if key in observed_country_keys:
            issues.append(f"{row_label} duplicates the {country}/{field} slot.")
            continue
        observed_country_keys.add(key)

        entry, entry_issues = validate_entry(
            row,
            row_label,
            provider_map,
            COUNTRY_FIELD_LOOKUP[field],
            "country",
            issue_recorder,
            today,
        )
        issues.extend(entry_issues)
        if entry is None:
            continue
        country_values.setdefault(country, {})[field] = entry
        filled_country_rows += 1

    for row_index, row in enumerate(city_rows, start=2):
        city_id = row.get("city_id", "").strip()
        field = row.get("field", "").strip()
        row_label = f"city_inputs.csv row {row_index} ({city_id or 'unknown'} / {field or 'unknown'})"
        if city_id not in city_ids:
            issues.append(f"{row_label} references unknown city_id '{city_id}'.")
            continue
        if field not in CITY_FIELD_LOOKUP:
            issues.append(f"{row_label} references unknown field '{field}'.")
            continue

        key = (city_id, field)
        if key in observed_city_keys:
            issues.append(f"{row_label} duplicates the {city_id}/{field} slot.")
            continue
        observed_city_keys.add(key)

        entry, entry_issues = validate_entry(
            row,
            row_label,
            provider_map,
            CITY_FIELD_LOOKUP[field],
            "city",
            issue_recorder,
            today,
        )
        issues.extend(entry_issues)
        if entry is None:
            continue
        city_values.setdefault(city_id, {})[field] = entry
        filled_city_rows += 1

    missing_country_keys = expected_country_keys - observed_country_keys
    missing_city_keys = expected_city_keys - observed_city_keys
    if missing_country_keys:
        issues.append(
            f"country_context.csv is missing {len(missing_country_keys)} expected country-field rows."
        )
    if missing_city_keys:
        issues.append(f"city_inputs.csv is missing {len(missing_city_keys)} expected city-field rows.")

    integrity_summaries = issue_recorder.summaries()
    issues.extend(summarize_integrity_issues(integrity_summaries))

    stats = {
        "country_slots": len(expected_country_keys),
        "country_values": filled_country_rows,
        "city_slots": len(expected_city_keys),
        "city_values": filled_city_rows,
        "country_count": len(countries),
        "city_count": len(city_ids),
        "integrity_issue_count": sum(summary.count for summary in integrity_summaries),
    }

    return SourcePackValidation(
        providers=provider_map,
        country_values=country_values,
        city_values=city_values,
        issues=issues,
        integrity_summaries=integrity_summaries,
        stats=stats,
    )


def percentile_inc(values: list[float], percentile: float) -> float | None:
    if len(values) < 2:
        return None

    ordered = sorted(values)
    index = (len(ordered) - 1) * percentile
    lower = int(index)
    upper = min(lower + 1, len(ordered) - 1)
    weight = index - lower
    return ordered[lower] + (ordered[upper] - ordered[lower]) * weight


def normalize_value(value: float | None, p05: float | None, p95: float | None, directionality: str) -> float | None:
    if value is None or p05 is None or p95 is None or p95 == p05:
        return None

    clamped = min(max(value, p05), p95)
    if directionality == "positive":
        normalized = 100 * (clamped - p05) / (p95 - p05)
    else:
        normalized = 100 * (p95 - clamped) / (p95 - p05)
    return max(0.0, min(100.0, normalized))


def round_score(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 1)


def summarise_sources(entries: list[SourceEntry]) -> dict[str, str]:
    if not entries:
        return {
            "source_url": "",
            "source_tier": "",
            "source_date": "",
            "reference_period": "",
            "source_scope": "",
            "proxy_status": "",
            "notes": "",
        }

    seen_urls: list[str] = []
    seen_tiers: list[str] = []
    seen_providers: list[str] = []
    seen_fields: list[str] = []
    seen_scopes: list[str] = []
    seen_proxy_statuses: list[str] = []
    seen_reference_periods: list[str] = []
    latest_date = max(entry.source_date for entry in entries)

    for entry in entries:
        if entry.source_url not in seen_urls:
            seen_urls.append(entry.source_url)
        if entry.tier not in seen_tiers:
            seen_tiers.append(entry.tier)
        if entry.provider_id not in seen_providers:
            seen_providers.append(entry.provider_id)
        if entry.field not in seen_fields:
            seen_fields.append(entry.field)
        if entry.source_scope not in seen_scopes:
            seen_scopes.append(entry.source_scope)
        if entry.proxy_status not in seen_proxy_statuses:
            seen_proxy_statuses.append(entry.proxy_status)
        if entry.reference_period not in seen_reference_periods:
            seen_reference_periods.append(entry.reference_period)

    seen_tiers.sort(key=lambda tier: TIER_ORDER.get(tier, 99))
    compact_urls = seen_urls[:3]
    if len(seen_urls) > 3:
        compact_urls.append(f"+{len(seen_urls) - 3} more")

    note_parts = [
        f"providers={', '.join(seen_providers[:4])}",
        f"fields={', '.join(seen_fields[:4])}",
        f"scope={', '.join(seen_scopes[:3])}",
        f"proxy={', '.join(seen_proxy_statuses[:3])}",
    ]
    if len(seen_fields) > 4:
        note_parts.append(f"+{len(seen_fields) - 4} additional fields")

    return {
        "source_url": " | ".join(compact_urls),
        "source_tier": ", ".join(seen_tiers),
        "source_date": latest_date,
        "reference_period": " | ".join(seen_reference_periods[:3]),
        "source_scope": ", ".join(seen_scopes[:3]),
        "proxy_status": ", ".join(seen_proxy_statuses[:3]),
        "notes": "; ".join(note_parts),
    }


def apply_source_pack_to_workbook(validation: SourcePackValidation) -> dict[str, int]:
    workbook = load_workbook(WORKBOOK_PATH)
    country_sheet = workbook["Country_Context"]
    city_sheet = workbook["City_Inputs"]

    country_headers = {country_sheet.cell(row=1, column=index).value: index for index in range(1, country_sheet.max_column + 1)}
    city_headers = {city_sheet.cell(row=1, column=index).value: index for index in range(1, city_sheet.max_column + 1)}

    country_row_map = {
        str(country_sheet.cell(row=row_index, column=country_headers["country"]).value): row_index
        for row_index in range(2, country_sheet.max_row + 1)
    }
    city_row_map = {
        str(city_sheet.cell(row=row_index, column=city_headers["city_id"]).value): row_index
        for row_index in range(2, city_sheet.max_row + 1)
    }

    written_country_values = 0
    written_city_values = 0

    for country, row_index in country_row_map.items():
        field_entries = validation.country_values.get(country, {})
        for field in COUNTRY_FIELD_ORDER:
            value = field_entries[field].value if field in field_entries else ""
            country_sheet.cell(row=row_index, column=country_headers[field], value=value)
            if value != "":
                written_country_values += 1

        summary = summarise_sources(list(field_entries.values()))
        country_sheet.cell(row=row_index, column=country_headers["source_url"], value=summary["source_url"])
        country_sheet.cell(row=row_index, column=country_headers["source_tier"], value=summary["source_tier"] or "")
        country_sheet.cell(row=row_index, column=country_headers["update_date"], value=summary["source_date"] or "")
        if "reference_period" in country_headers:
            country_sheet.cell(row=row_index, column=country_headers["reference_period"], value=summary["reference_period"] or "")
        if "source_scope" in country_headers:
            country_sheet.cell(row=row_index, column=country_headers["source_scope"], value=summary["source_scope"] or "")
        if "proxy_status" in country_headers:
            country_sheet.cell(row=row_index, column=country_headers["proxy_status"], value=summary["proxy_status"] or "")
        country_sheet.cell(
            row=row_index,
            column=country_headers["notes"],
            value=summary["notes"] or "Populate official country context inputs before ranking.",
        )

    for city_id, row_index in city_row_map.items():
        field_entries = validation.city_values.get(city_id, {})
        for field in CITY_FIELD_ORDER:
            value = field_entries[field].value if field in field_entries else ""
            city_sheet.cell(row=row_index, column=city_headers[field], value=value)
            if value != "":
                written_city_values += 1

        country = str(city_sheet.cell(row=row_index, column=city_headers["country"]).value)
        country_entries = validation.country_values.get(country, {})
        city_summary = summarise_sources(
            list(field_entries.values()) + list(country_entries.values())
        )
        if "reference_period" in city_headers:
            city_sheet.cell(row=row_index, column=city_headers["reference_period"], value=city_summary["reference_period"])
        if "source_scope" in city_headers:
            city_sheet.cell(row=row_index, column=city_headers["source_scope"], value=city_summary["source_scope"])
        if "proxy_status" in city_headers:
            city_sheet.cell(row=row_index, column=city_headers["proxy_status"], value=city_summary["proxy_status"])
        for pillar in PILLAR_ORDER:
            entries = [
                field_entries[field]
                for field in CITY_FIELDS_BY_PILLAR[pillar]
                if field in field_entries
            ]
            entries.extend(
                country_entries[field]
                for field in COUNTRY_FIELDS_BY_PILLAR[pillar]
                if field in country_entries
            )
            summary = summarise_sources(entries)
            city_sheet.cell(row=row_index, column=city_headers[f"{pillar}_source_url"], value=summary["source_url"])
            city_sheet.cell(row=row_index, column=city_headers[f"{pillar}_source_tier"], value=summary["source_tier"])
            city_sheet.cell(row=row_index, column=city_headers[f"{pillar}_source_date"], value=summary["source_date"])
            city_sheet.cell(row=row_index, column=city_headers[f"{pillar}_notes"], value=summary["notes"])

    workbook.save(WORKBOOK_PATH)
    return {"country_values_written": written_country_values, "city_values_written": written_city_values}


def city_row_from_sources(city: dict[str, str], validation: SourcePackValidation) -> dict[str, Any]:
    row: dict[str, Any] = {
        "city_id": city["city_id"],
        "display_name": city["display_name"],
        "country": city["country"],
        "cohort": city["cohort"],
        "city_type": city["city_type"],
        "manifest_status": city["manifest_status"],
        "inclusion_rationale": city["inclusion_rationale"],
    }

    city_entries = validation.city_values.get(city["city_id"], {})
    country_entries = validation.country_values.get(city["country"], {})

    for field in CITY_FIELD_ORDER:
        row[field] = city_entries[field].value if field in city_entries else None

    for field, context_field in COUNTRY_CONTEXT_FIELD_MAP.items():
        value = country_entries[field].value if field in country_entries else None
        row[context_field] = value

    gross_income = row["gross_income"]
    rent = row["rent"]
    utilities = row["utilities"]
    transit_cost = row["transit_cost"]
    internet_cost = row["internet_cost"]
    food_cost = row["food_cost"]
    ppp_factor = row["ppp_private_consumption_context"]
    tax_rate = row["tax_rate_context"]
    if None not in (gross_income, rent, utilities, transit_cost, internet_cost, food_cost, ppp_factor, tax_rate):
        row["di_ppp_raw"] = ((gross_income * (1 - tax_rate)) - rent - utilities - transit_cost - internet_cost - food_cost) / ppp_factor
    else:
        row["di_ppp_raw"] = None

    row["household_debt_effective_raw"] = (
        row["household_debt_burden_raw"]
        if row["household_debt_burden_raw"] is not None
        else row["household_debt_proxy_context"]
    )
    return row


def compute_ranked_rows(validation: SourcePackValidation) -> list[dict[str, Any]]:
    city_universe = build_city_universe()
    city_rows = [city_row_from_sources(city, validation) for city in city_universe]

    norm_stats: dict[str, tuple[float | None, float | None]] = {}
    for input_key, spec in SCORE_INPUTS.items():
        values = [
            float(city_row[spec["source_column"]])
            for city_row in city_rows
            if city_row[spec["source_column"]] is not None
        ]
        norm_stats[input_key] = (percentile_inc(values, 0.05), percentile_inc(values, 0.95))

    thresholds = {
        "ranked_min_overall": 0.5,
        "ranked_min_pillar": 0.35,
        "grade_a_min": 0.75,
        "grade_b_min": 0.5,
        "grade_c_min": 0.35,
    }

    scored_rows: list[dict[str, Any]] = []
    for city_row in city_rows:
        metric_scores: dict[str, float | None] = {}
        metric_coverage: dict[str, float | None] = {}

        for metric_key, spec in METRIC_SPECS.items():
            if spec["type"] == "direct":
                input_key = str(spec["input_key"])
                p05, p95 = norm_stats[input_key]
                directionality = SCORE_INPUTS[input_key]["directionality"]
                score = normalize_value(city_row[input_key], p05, p95, directionality)
                metric_scores[metric_key] = score
                metric_coverage[metric_key] = 1.0 if score is not None else None
                continue

            numerator = 0.0
            denominator = 0.0
            availability = 0.0
            for input_key, weight in spec["components"]:
                p05, p95 = norm_stats[input_key]
                directionality = SCORE_INPUTS[input_key]["directionality"]
                component_score = normalize_value(city_row[input_key], p05, p95, directionality)
                if component_score is None:
                    continue
                numerator += component_score * float(weight)
                denominator += float(weight)
                availability += float(weight)
            metric_scores[metric_key] = numerator / denominator if denominator else None
            metric_coverage[metric_key] = availability if availability else None

        pillar_scores: dict[str, float | None] = {}
        pillar_coverage: dict[str, float | None] = {}
        for pillar, metrics in PILLAR_METRICS.items():
            numerator = 0.0
            denominator = 0.0
            coverage_numerator = 0.0
            total_weight = sum(weight for _, weight in metrics)
            for metric_key, weight in metrics:
                metric_score = metric_scores[metric_key]
                if metric_score is not None:
                    numerator += metric_score * weight
                    denominator += weight
                coverage_value = metric_coverage[metric_key]
                if coverage_value is not None:
                    coverage_numerator += coverage_value * weight
            pillar_scores[pillar] = numerator / denominator if denominator else None
            pillar_coverage[pillar] = coverage_numerator / total_weight if coverage_numerator else None

        if all(pillar_coverage[pillar] is None for pillar in PILLAR_ORDER):
            overall_coverage = None
        else:
            overall_coverage = sum((pillar_coverage[pillar] or 0.0) * PILLAR_WEIGHTS[pillar] for pillar in PILLAR_ORDER) / 100

        if overall_coverage is None:
            coverage_grade = ""
        elif overall_coverage >= thresholds["grade_a_min"]:
            coverage_grade = "A"
        elif overall_coverage >= thresholds["grade_b_min"]:
            coverage_grade = "B"
        elif overall_coverage >= thresholds["grade_c_min"]:
            coverage_grade = "C"
        else:
            coverage_grade = "Watchlist"

        ranking_status = "Ranked"
        if overall_coverage is None or overall_coverage < thresholds["ranked_min_overall"]:
            ranking_status = "Watchlist"
        for pillar in PILLAR_ORDER:
            coverage_value = pillar_coverage[pillar]
            if coverage_value is None or coverage_value < thresholds["ranked_min_pillar"]:
                ranking_status = "Watchlist"
                break

        available_weight = sum(
            PILLAR_WEIGHTS[pillar] for pillar in PILLAR_ORDER if pillar_scores[pillar] is not None
        )
        slic_score = (
            sum((pillar_scores[pillar] or 0.0) * PILLAR_WEIGHTS[pillar] for pillar in PILLAR_ORDER) / available_weight
            if available_weight
            else None
        )

        scored_rows.append(
            {
                "cityId": city_row["city_id"],
                "displayName": city_row["display_name"],
                "country": city_row["country"],
                "region": city_row["cohort"],
                "manifestStatus": city_row["manifest_status"],
                "cityType": city_row["city_type"],
                "coverageGrade": coverage_grade,
                "overallWeightedCoverage": round_score(overall_coverage),
                "pressureScore": round_score(pillar_scores["pressure"]),
                "viabilityScore": round_score(pillar_scores["viability"]),
                "capabilityScore": round_score(pillar_scores["capability"]),
                "communityScore": round_score(pillar_scores["community"]),
                "creativeScore": round_score(pillar_scores["creative"]),
                "slicScore": round_score(slic_score),
                "rankingStatus": ranking_status,
            }
        )

    ranked_rows = [
        row for row in scored_rows if row["rankingStatus"] == "Ranked" and row["slicScore"] is not None
    ]

    for row in ranked_rows:
        row["rank"] = 1 + sum(
            1
            for other in ranked_rows
            if other["slicScore"] is not None and row["slicScore"] is not None and other["slicScore"] > row["slicScore"]
        )

    ranked_rows.sort(
        key=lambda row: (
            int(row["rank"]),
            -(row["slicScore"] or 0.0),
            row["displayName"],
        )
    )
    return ranked_rows


def build_integrity_dashboard_rows(validation: SourcePackValidation) -> list[dict[str, str]]:
    rows = [
        {
            "section": "Coverage",
            "metric": "Country valid rows",
            "value": f"{validation.stats['country_values']}/{validation.stats['country_slots']}",
            "detail": "Rows that passed schema, provider, freshness, scope, and proxy checks.",
        },
        {
            "section": "Coverage",
            "metric": "City valid rows",
            "value": f"{validation.stats['city_values']}/{validation.stats['city_slots']}",
            "detail": "Rows that passed schema, provider, freshness, scope, and proxy checks.",
        },
        {
            "section": "Integrity",
            "metric": "Integrity findings",
            "value": str(validation.stats.get("integrity_issue_count", 0)),
            "detail": "Aggregated validation findings across the source pack.",
        },
    ]
    rows.extend(
        {
            "section": "Integrity",
            "metric": summary.label,
            "value": str(summary.count),
            "detail": " | ".join(summary.samples),
        }
        for summary in validation.integrity_summaries
    )
    return rows


def build_city_integrity_watchlist(validation: SourcePackValidation) -> list[dict[str, str]]:
    city_universe = build_city_universe()
    scope_order = {scope: index for index, scope in enumerate(SOURCE_SCOPE_VALUES)}
    required_country_fields = [spec for spec in COUNTRY_FIELD_SPECS if spec["required_for_ranking"] == "yes"]
    required_city_fields = [spec for spec in CITY_FIELD_SPECS if spec["required_for_ranking"] == "yes"]
    rows: list[dict[str, str]] = []

    for city in city_universe:
        city_entries = validation.city_values.get(city["city_id"], {})
        country_entries = validation.country_values.get(city["country"], {})
        missing_city_fields = [spec["field"] for spec in required_city_fields if spec["field"] not in city_entries]
        missing_country_fields = [spec["field"] for spec in required_country_fields if spec["field"] not in country_entries]
        required_entries = [
            city_entries[spec["field"]]
            for spec in required_city_fields
            if spec["field"] in city_entries
        ]
        required_entries.extend(
            country_entries[spec["field"]]
            for spec in required_country_fields
            if spec["field"] in country_entries
        )

        broadest_scope = ""
        if required_entries:
            broadest_scope = max(
                (entry.source_scope for entry in required_entries),
                key=lambda scope: scope_order.get(scope, -1),
            )

        proxy_count = sum(1 for entry in required_entries if entry.proxy_status == "approved_proxy")
        qualified = not missing_city_fields and not missing_country_fields
        next_action = "Ready for publication gate review."
        if missing_country_fields:
            next_action = "Complete missing country context rows first."
        elif missing_city_fields:
            next_action = "Complete missing city input rows."
        elif proxy_count:
            next_action = "Review approved proxies and replace with more local direct evidence where possible."

        rows.append(
            {
                "city_id": city["city_id"],
                "display_name": city["display_name"],
                "country": city["country"],
                "qualification_status": "Qualified" if qualified else "Needs evidence",
                "valid_required_city_fields": f"{len(required_city_fields) - len(missing_city_fields)}/{len(required_city_fields)}",
                "valid_required_country_fields": f"{len(required_country_fields) - len(missing_country_fields)}/{len(required_country_fields)}",
                "missing_city_fields": ", ".join(missing_city_fields[:8]),
                "missing_country_fields": ", ".join(missing_country_fields[:6]),
                "broadest_source_scope": broadest_scope,
                "approved_proxy_rows": str(proxy_count),
                "next_priority_action": next_action,
            }
        )

    rows.sort(
        key=lambda row: (
            0 if row["qualification_status"] != "Qualified" else 1,
            row["country"],
            row["display_name"],
        )
    )
    return rows


def source_pack_completion_summary(validation: SourcePackValidation) -> str:
    return (
        f"country values={validation.stats['country_values']}/{validation.stats['country_slots']}, "
        f"city values={validation.stats['city_values']}/{validation.stats['city_slots']}, "
        f"integrity findings={validation.stats.get('integrity_issue_count', 0)}"
    )
